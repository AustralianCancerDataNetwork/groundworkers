"""XLSX decomposer — produces one ``RawTable`` per sheet using openpyxl."""

from __future__ import annotations

import io
from typing import Any

from groundworkers.services.source_planning.models import RawTable, SourceFormat

_MAX_SAMPLE = 5


def decompose(content: bytes, filename: str | None = None) -> list[RawTable]:
    """Open XLSX bytes and return one ``RawTable`` per worksheet."""

    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    except Exception:
        return []

    sheet_count = len(wb.worksheets)
    tables: list[RawTable] = []
    for sheet_index, ws in enumerate(wb.worksheets):
        raw_rows = _read_sheet(ws)
        if not raw_rows:
            continue

        banner_row_removed = False
        if _is_banner_row(raw_rows):
            raw_rows = raw_rows[1:]
            banner_row_removed = True
        if not raw_rows:
            continue

        headers = [str(cell) if cell is not None else "" for cell in raw_rows[0]]
        headers = _dedupe_headers(headers)
        data_rows = raw_rows[1:]
        rows: list[dict[str, str]] = []
        for raw_row in data_rows:
            row_dict: dict[str, str] = {}
            for index, header in enumerate(headers):
                value = raw_row[index] if index < len(raw_row) else None
                row_dict[header] = str(value) if value is not None else ""
            rows.append(row_dict)

        tables.append(
            RawTable(
                name=ws.title or f"Sheet{sheet_index + 1}",
                headers=headers,
                rows=rows,
                sample_rows=rows[:_MAX_SAMPLE],
                source_format=SourceFormat.XLSX,
                row_count=len(rows),
                metadata={
                    "sheet_index": sheet_index,
                    "sheet_count": sheet_count,
                    "banner_row_removed": banner_row_removed,
                },
            )
        )

    wb.close()
    return tables


def _build_merge_fill(ws) -> dict[tuple[int, int], Any]:
    fill: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                if row_idx == merged_range.min_row and col_idx == merged_range.min_col:
                    continue
                fill[(row_idx, col_idx)] = top_left_value
    return fill


def _read_sheet(ws) -> list[list]:
    merge_fill = _build_merge_fill(ws)
    rows = []
    for row in ws.iter_rows():
        row_values = []
        for cell in row:
            coord = (cell.row, cell.column)
            row_values.append(merge_fill[coord] if coord in merge_fill else cell.value)
        rows.append(row_values)
    return rows


def _is_banner_row(rows: list[list]) -> bool:
    if len(rows) < 2:
        return False
    first_populated = sum(1 for cell in rows[0] if cell is not None and str(cell).strip())
    second_populated = sum(1 for cell in rows[1] if cell is not None and str(cell).strip())
    return first_populated <= 1 and second_populated > first_populated


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for header in headers:
        if not header:
            header = "col"
        if header in seen:
            seen[header] += 1
            result.append(f"{header}_{seen[header]}")
        else:
            seen[header] = 1
            result.append(header)
    return result
